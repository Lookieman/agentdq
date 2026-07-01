# v0.1 | 27-Jun-2026 | Initial SAP SE16N xlsx extract loader
# v0.2 | 27-Jun-2026 | Rename key_field to header_anchor to reflect its true role
#                      (locating the header row, not defining the business key)
# v0.3 | 27-Jun-2026 | Renamed module from src/data/loader.py to extract_loader.py
#                      to avoid clashing with the rules loader

"""Loader for SAP table extracts exported from SE16N as Excel workbooks.

The SE16N Excel export is not a plain table. It carries a short preamble and a
spacer column that must be stripped before the data is usable. A MARA export
looks like this:

    row 1            : blank
    row 2            : 'Table:'            <table name>
    row 3            : 'Displayed Fields:' <counts>
    row 4            : blank
    row 5            : column headers (field names)
    row 6            : blank separator
    row 7 onward     : data
    column 0         : an empty spacer column throughout

The header_anchor argument is the field name used only to locate that header
row within the preamble; it has nothing to do with the table's business key,
which is a composite handled in the schema and profiling layers.

Two SAP-specific concerns are handled deliberately:

- Key fields such as MATNR are exported with leading zeros
  (for example '000000000000000011') and must remain strings; every column is
  read as text so no leading zero is lost to numeric coercion.
- An empty cell means the field is unpopulated; it is normalised to None so
  downstream population checks behave correctly.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd


def _to_text(cell: object) -> Optional[str]:
    """Coerce a cell value to text, preserving leading zeros.

    Empty strings and None both become None so that population checks treat
    them as unpopulated. Integral floats (an occasional artefact of Excel
    storage) are rendered without a trailing '.0'.
    """
    text: Optional[str] = None

    if cell is None:
        text = None
    elif isinstance(cell, str):
        text = cell if cell != "" else None
    elif isinstance(cell, float) and cell.is_integer():
        text = str(int(cell))
    else:
        text = str(cell)

    return text


def _find_header_row(worksheet: object, header_anchor: str, max_scan: int) -> Optional[int]:  # v0.2
    """Return the 1-indexed row number that holds the column headers.

    The header row is the first row within the scan window that contains the
    anchor field (MATNR for the material tables). Returning None lets the
    caller raise a clear error rather than silently guessing.
    """
    header_row: Optional[int] = None
    row_index: int = 0

    for row in worksheet.iter_rows(min_row=1, max_row=max_scan, values_only=True):
        row_index += 1
        if any(cell == header_anchor for cell in row):  # v0.2
            header_row = row_index
            break

    return header_row


def load_sap_table(
    path: str,
    header_anchor: str = "MATNR",  # v0.2
    sheet: str = "Data",
    max_header_scan: int = 25,
) -> pd.DataFrame:
    """Load a single SE16N xlsx extract into a tidy DataFrame.

    header_anchor names the field used to locate the header row only; it does
    not define uniqueness. All columns are strings, leading zeros are
    preserved, the spacer column and preamble are dropped, and empty cells
    become None. Rows that are empty across every mapped column (such as the
    blank separator) are skipped.
    """
    file_path: Path = Path(path)
    workbook: object = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    worksheet: object = None
    header_row: Optional[int] = None
    header_cells: tuple = ()
    column_map: dict[int, str] = {}
    records: list[dict[str, Optional[str]]] = []
    frame: pd.DataFrame = None

    if sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    header_row = _find_header_row(worksheet, header_anchor, max_header_scan)  # v0.2
    if header_row is None:
        raise ValueError(
            f"could not locate a header row containing '{header_anchor}' in "  # v0.2
            f"{file_path.name}; check the export format or pass a different header_anchor"  # v0.2
        )

    # Map column index -> field name for non-empty header cells. This drops the
    # leading spacer column and any other blank header positions.
    header_cells = list(
        worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
    )[0]
    for idx, name in enumerate(header_cells):
        if name not in (None, ""):
            column_map[idx] = str(name).strip()

    # Data begins below the header. The single blank separator row is handled by
    # skipping any row that is empty across all mapped columns.
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        values: dict[str, Optional[str]] = {}
        is_empty: bool = True
        for idx, field_name in column_map.items():
            cell = row[idx] if idx < len(row) else None
            text = _to_text(cell)
            values[field_name] = text
            if text is not None:
                is_empty = False
        if not is_empty:
            records.append(values)

    workbook.close()
    frame = pd.DataFrame.from_records(records, columns=list(column_map.values()))
    return frame
