# v0.1 | 27-Jun-2026 | Initial Information Steward rule importer
# v0.2 | 27-Jun-2026 | Clear error when the Consolidated sheet is missing (wrong workbook)
# v0.3 | 27-Jun-2026 | Defer multi-parameter expressions as cross_field before the IN test

"""Importer that turns the Information Steward workbook into canonical RuleSpec IR.

It reads the Consolidated sheet, classifies each rule into an archetype, builds
the declarative assertion predicate, re-maps the legacy dimension onto a DAMA
dimension, and binds the rule against the current schema to decide whether it can
run on the synthetic data we generate. Provenance (original name, expression,
reference table and documentation) travels with every rule.

The three deterministic archetypes are handled in full:

- not_null          -> assertion: field IS_NOT_NULL
- domain_in         -> assertion: field IN [inline values from the expression]
- reference_exists  -> assertion: field IN [domain resolved from our schema]

Compositional rules (cross-field boolean logic, regex within a condition) are not
auto-parsed in this first cut; they are listed in the import report for manual
formalisation, with curated examples authored separately to exercise the IR.

reference_exists rules bind their permitted values from our self-defined schema
domains rather than the original SAP check tables, in line with the synthetic
benchmark approach. A rule whose field is absent from the schema, or whose domain
cannot be resolved, is retained but marked not executable.

Run as a module:

    python -m src.rules.is_importer --workbook Infomation_Steward_Rules_Tables.xlsx \\
        --schema config/schema --tables MARA,MARC,MAKT --out config/rules
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Optional

import openpyxl
import yaml

from src.contracts import (
    Comparison,
    Dimension,
    ISDimension,
    Operator,
    Provenance,
    RuleArchetype,
    RuleSpec,
    Severity,
    map_to_dama,
)
from src.data.schema import TableSchema, load_schemas


SHEET_NAME: str = "Consolidated"

# Column indices within the Consolidated sheet.
COL_OBJNAME: int = 15
COL_NEW_OBJNAME: int = 16
COL_RULE_DOC: int = 17
COL_DIMENSION: int = 18
COL_TABNAME: int = 21
COL_FIELDNAME: int = 22
COL_LOOKUP_TAB: int = 24
COL_ISRULE: int = 26

IN_LIST_RE: re.Pattern = re.compile(r"\bIN\b\s*\(([^)]*)\)", re.IGNORECASE)
QUOTED_RE: re.Pattern = re.compile(r"'([^']*)'")
PARAM_RE: re.Pattern = re.compile(r"\$P_I_\w+")  # v0.3


def _is_dimension_from_str(text: Optional[str]) -> Optional[ISDimension]:
    """Map a dimension cell to an ISDimension, tolerating case and blanks."""
    cleaned: str = ""
    member: ISDimension = None

    if not text:
        return None
    cleaned = str(text).strip().lower()
    for member in ISDimension:
        if member.value.lower() == cleaned:
            return member
    return None


def _classify(expression: Optional[str], lookup_tab: Optional[str]) -> RuleArchetype:
    """Classify a rule expression into an archetype.

    A rule that references more than one parameter is compositional (cross-field)
    and is recognised as such before the single-field tests, so a cross-field
    rule that merely contains an IN clause is not mistaken for a domain check.
    """
    text: str = expression or ""
    params: set[str] = set(PARAM_RE.findall(text))  # v0.3

    if len(params) > 1:  # v0.3
        return RuleArchetype.CROSS_FIELD  # v0.3
    if "exists(" in text or (lookup_tab not in (None, "", ".")):
        return RuleArchetype.REFERENCE_EXISTS
    if IN_LIST_RE.search(text):
        return RuleArchetype.DOMAIN_IN
    if "IS NOT NULL" in text.upper():
        return RuleArchetype.NOT_NULL
    if "match_regex" in text or "match_pattern" in text:
        return RuleArchetype.FORMAT_REGEX
    if text.strip():
        return RuleArchetype.CROSS_FIELD
    return RuleArchetype.OTHER


def _parse_domain_in(expression: str) -> list[str]:
    """Extract the inline value list from an IN (...) expression."""
    match: Optional[re.Match] = IN_LIST_RE.search(expression)
    values: list[str] = []

    if match is None:
        return values
    values = QUOTED_RE.findall(match.group(1))
    return values


def _single_field(fieldname: Optional[str]) -> Optional[str]:
    """Return the single bound field, or None when the cell names several."""
    text: str = ""
    tokens: list[str] = []

    if not fieldname:
        return None
    text = str(fieldname).replace(",", " ").replace("\n", " ").strip()
    tokens = [t for t in text.split() if t]
    if len(tokens) == 1:
        return tokens[0]
    return None


def _rule_name(objname: Optional[str], new_objname: Optional[str]) -> str:
    """Prefer the new object name, falling back to the original."""
    if new_objname:
        return str(new_objname).strip()
    if objname:
        return str(objname).strip()
    return "UNNAMED_RULE"


def _build_rule(
    row: tuple,
    schema: Optional[TableSchema],
) -> tuple[Optional[RuleSpec], Optional[str]]:
    """Build one RuleSpec from a sheet row.

    Returns (rule, None) on success or (None, reason) when the rule is deferred
    for manual formalisation.
    """
    table: str = str(row[COL_TABNAME])
    field: Optional[str] = _single_field(row[COL_FIELDNAME])
    expression: str = str(row[COL_ISRULE] or "")
    lookup_tab: Optional[str] = row[COL_LOOKUP_TAB]
    rule_doc: Optional[str] = row[COL_RULE_DOC]
    name: str = _rule_name(row[COL_OBJNAME], row[COL_NEW_OBJNAME])
    is_dim: Optional[ISDimension] = _is_dimension_from_str(row[COL_DIMENSION])
    archetype: RuleArchetype = _classify(expression, lookup_tab)
    assertion: Optional[Comparison] = None
    domain_values: list[str] = []
    executable: bool = True
    binding_notes: Optional[str] = None
    field_in_schema: bool = False
    severity: Severity = Severity.MEDIUM

    # Compositional rules are deferred in this first cut.
    if archetype in (RuleArchetype.CROSS_FIELD, RuleArchetype.FORMAT_REGEX, RuleArchetype.OTHER):
        return None, f"{archetype.value} expression not auto-parsed in first cut"
    if field is None:
        return None, "multi-field or unparseable field binding"

    field_in_schema = schema is not None and schema.field(field) is not None

    if archetype == RuleArchetype.NOT_NULL:
        assertion = Comparison(field=field, op=Operator.IS_NOT_NULL)
        if not field_in_schema:
            executable = False
            binding_notes = "field not present in current schema projection"
        elif schema.field(field).mandatory:
            severity = Severity.HIGH

    elif archetype == RuleArchetype.DOMAIN_IN:
        domain_values = _parse_domain_in(expression)
        assertion = Comparison(field=field, op=Operator.IN, value=domain_values)
        if not domain_values:
            executable = False
            binding_notes = "could not parse inline value list"
        elif not field_in_schema:
            executable = False
            binding_notes = "field not present in current schema projection"

    elif archetype == RuleArchetype.REFERENCE_EXISTS:
        # Synthetic path: bind permitted values from our self-defined domain.
        if field_in_schema and schema.domain(field):
            domain_values = schema.domain(field)
            assertion = Comparison(field=field, op=Operator.IN, value=domain_values)
        else:
            assertion = Comparison(field=field, op=Operator.IN, value=None)
            executable = False
            binding_notes = "no self-defined domain for field; resolve from check table on the real-data path"

    if assertion is None:
        return None, "no assertion could be built"

    dama: Dimension = map_to_dama(is_dim, archetype)
    provenance: Provenance = Provenance(
        source="information_steward",
        original_name=name,
        original_expression=expression.strip() if expression else None,
        reference_table=str(lookup_tab) if lookup_tab not in (None, "", ".") else None,
        rule_doc=str(rule_doc).strip() if rule_doc else None,
    )

    rule: RuleSpec = RuleSpec(
        rule_id=f"IS_{table}_{field}_{archetype.value}".upper(),
        name=name,
        table=table,
        dama_dimension=dama,
        is_dimension=is_dim,
        archetype=archetype,
        severity=severity,
        description=str(rule_doc).strip() if rule_doc else "",
        fields=[field],
        assertion=assertion,
        executable=executable,
        binding_notes=binding_notes,
        provenance=provenance,
    )
    return rule, None


def import_is_rules(
    workbook_path: str,
    tables: list[str],
    schema_dir: Optional[str],
    out_dir: Optional[str],
) -> dict[str, Any]:
    """Import rules for the given tables and optionally write rule YAMLs."""
    workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)  # v0.2
    worksheet = None  # v0.2
    schemas: dict[str, TableSchema] = {}
    rules_by_table: dict[str, list[RuleSpec]] = {}
    deferred: list[dict[str, str]] = []
    target: set[str] = set(tables)
    table: str = ""

    if SHEET_NAME not in workbook.sheetnames:  # v0.2
        workbook.close()  # v0.2
        raise ValueError(  # v0.2
            f"'{SHEET_NAME}' sheet not found in {Path(workbook_path).name} "  # v0.2
            f"(found: {', '.join(workbook.sheetnames)}). This importer expects the "  # v0.2
            f"Information Steward rules workbook, not a data extract."  # v0.2
        )  # v0.2
    worksheet = workbook[SHEET_NAME]  # v0.2
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))  # v0.2

    if schema_dir is not None:
        schemas = load_schemas(schema_dir, tables)
    for table in tables:
        rules_by_table[table] = []

    for row in rows:
        if row[COL_TABNAME] not in target:
            continue
        table = str(row[COL_TABNAME])
        rule, reason = _build_rule(row, schemas.get(table))
        if rule is not None:
            rules_by_table[table].append(rule)
        elif reason is not None:
            deferred.append(
                {
                    "table": table,
                    "name": _rule_name(row[COL_OBJNAME], row[COL_NEW_OBJNAME]),
                    "reason": reason,
                }
            )

    workbook.close()
    report: dict[str, Any] = _summarise(rules_by_table, deferred)

    if out_dir is not None:
        _write_outputs(rules_by_table, report, Path(out_dir))

    return report


def _summarise(
    rules_by_table: dict[str, list[RuleSpec]],
    deferred: list[dict[str, str]],
) -> dict[str, Any]:
    """Build a structured import report."""
    report: dict[str, Any] = {"tables": {}, "deferred_count": len(deferred), "deferred": deferred}
    table: str = ""
    rules: list[RuleSpec] = []
    rule: RuleSpec = None
    by_archetype: dict[str, int] = {}
    by_dimension: dict[str, int] = {}
    executable_count: int = 0

    for table, rules in rules_by_table.items():
        by_archetype = {}
        by_dimension = {}
        executable_count = 0
        for rule in rules:
            by_archetype[rule.archetype.value] = by_archetype.get(rule.archetype.value, 0) + 1
            by_dimension[rule.dama_dimension.value] = by_dimension.get(rule.dama_dimension.value, 0) + 1
            if rule.executable:
                executable_count += 1
        report["tables"][table] = {
            "imported": len(rules),
            "executable": executable_count,
            "dormant": len(rules) - executable_count,
            "by_archetype": by_archetype,
            "by_dama_dimension": by_dimension,
        }

    return report


def _write_outputs(
    rules_by_table: dict[str, list[RuleSpec]],
    report: dict[str, Any],
    out_dir: Path,
) -> None:
    """Write one rule YAML per table plus a JSON import report."""
    out_dir.mkdir(parents=True, exist_ok=True)
    table: str = ""
    rules: list[RuleSpec] = []
    payload: dict[str, Any] = {}

    for table, rules in rules_by_table.items():
        payload = {
            "table": table,
            "rule_count": len(rules),
            "rules": [json.loads(rule.model_dump_json(exclude_none=True)) for rule in rules],
        }
        with (out_dir / f"{table.lower()}_rules.yaml").open("w", encoding="utf-8") as handle:
            yaml.safe_dump(payload, handle, sort_keys=False, allow_unicode=True, default_flow_style=False)
        print(f"wrote {out_dir / (table.lower() + '_rules.yaml')} ({len(rules)} rules)")

    with (out_dir / "_import_report.json").open("w", encoding="utf-8") as handle:
        json.dump(report, handle, indent=2)
    print(f"wrote {out_dir / '_import_report.json'}")


def main() -> None:
    """Entry point for module execution."""
    parser: argparse.ArgumentParser = argparse.ArgumentParser(
        description="Import Information Steward rules into canonical RuleSpec IR."
    )
    parser.add_argument("--workbook", required=True, help="path to the IS rules workbook")
    parser.add_argument("--schema", default=None, help="schema directory for domain binding")
    parser.add_argument("--tables", default="MARA,MARC,MAKT", help="comma separated tables")
    parser.add_argument("--out", default=None, help="output directory for rule YAMLs")
    args: argparse.Namespace = parser.parse_args()
    table_list: list[str] = [t.strip() for t in args.tables.split(",") if t.strip()]

    report: dict[str, Any] = import_is_rules(args.workbook, table_list, args.schema, args.out)
    print("\nImport summary:")
    for table, stats in report["tables"].items():
        print(f"  {table}: {stats['imported']} imported "
              f"({stats['executable']} executable, {stats['dormant']} dormant)")
    print(f"  deferred for manual formalisation: {report['deferred_count']}")


if __name__ == "__main__":
    main()
