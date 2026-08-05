# ---------------------------------------------------------------------------
# tools/onboard_object.py
# v1.0 | 13-Jul-2026 | Initial creation. The v1 onboarding scaffolder
#                      (design doc 10.3): deterministic, no LLM. Reads a raw
#                      SE16N/SE12 xlsx, detects the header row and candidate
#                      composite key, maps fields to known roles, reports
#                      reference-table readiness, and emits a DRAFT object pack
#                      with TODO markers for the steward to confirm.
# v1.1 | 13-Jul-2026 | Key detection: search all arities, prune supersets,
#                      rank by left-most columns - a coincidentally unique
#                      description column no longer masks the true composite
#                      key (caught by the EQKT test).
# v2.0 | 13-Jul-2026 | Emit a draft SCHEMA stub (config/schema/<table>.yaml)
#                      rather than a separate object pack. The schema already
#                      carried primary_key and header_anchor; a parallel config
#                      was a mistake. One file per table is what a steward
#                      should write to onboard an object.
# v2.1 | 04-Aug-2026 | Package 4a. The draft uniqueness block now matches
#                      schema v0.4 (blocking_keys, weighted compare_fields,
#                      methods, bands) so a scaffolded schema loads.
# ---------------------------------------------------------------------------
"""Onboarding scaffolder - agent-free by design.

The task decomposition (design doc 10.1) shows most of onboarding is
deterministic: header detection is a scan, key detection is arithmetic over
column combinations, role mapping is a lookup against the controlled
vocabulary. This tool does exactly that much and STOPS: the output is a draft
the steward edits and confirms, never a silently written config. The v2
judgement layer (table identity, key sanity-check) slots in later without
changing this flow.

The output is a DRAFT SCHEMA STUB: the onboarding fields (header_anchor,
primary_key, file_pattern, uniqueness) plus a skeleton field block per column,
with TODO markers. The steward fills in domains, mandatory flags and types -
the judgement a scan cannot supply - and confirms.

Run:
    python -m tools.onboard_object --file data/raw/EQUI_EX_DATA.xlsx --table EQUI
    python -m tools.onboard_object --file ... --table EQUI --out config/schema/equi.yaml
"""

from __future__ import annotations

import argparse
import itertools
import re
from pathlib import Path
from typing import Any, Optional

import openpyxl
import pandas as pd
import yaml

from src.data.extract_loader import load_sap_table
from src.data.profiler import profile_table  # v2.0
from src.rules.rule_bank import load_field_roles


# SAP field names: short, uppercase, alphanumeric with underscores.
_FIELDNAME_RE: re.Pattern = re.compile(r"^[A-Z][A-Z0-9_]{1,29}$")

# How far into the sheet the header may plausibly sit, and how many candidate
# columns the key search considers (left-most bias keeps this honest - SAP
# key fields lead the table).
MAX_HEADER_SCAN: int = 25
MAX_KEY_COLUMNS: int = 8
MAX_KEY_ARITY: int = 3


# ---------------------------------------------------------------------------
# Header detection (independent of extract_loader, which needs a known anchor)
# ---------------------------------------------------------------------------

def detect_header(xlsx_path: str | Path, sheet: str = "Data") -> tuple[int, list[str]]:
    """Find the header row without knowing an anchor: the first row within the
    scan window where at least three cells look like SAP field names. Returns
    (1-indexed row, field names in column order, blanks skipped)."""
    workbook = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
    worksheet = workbook[sheet] if sheet in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    row_index: int = 0
    fieldish: list[str] = []
    header_row: int = 0
    header_fields: list[str] = []

    for row in worksheet.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN, values_only=True):
        row_index += 1
        fieldish = [
            str(cell).strip() for cell in row
            if isinstance(cell, str) and _FIELDNAME_RE.match(str(cell).strip())
        ]
        if len(fieldish) >= 3:
            header_row = row_index
            header_fields = fieldish
            break

    workbook.close()
    if header_row == 0:
        raise ValueError(
            f"could not locate a header row in {Path(xlsx_path).name}; "
            f"is this an SE16N/SE12 export?"
        )
    return header_row, header_fields


# ---------------------------------------------------------------------------
# Key detection: arithmetic, not AI (design doc 10.1)
# ---------------------------------------------------------------------------

def detect_key_candidates(frame: pd.DataFrame) -> list[list[str]]:
    """Search column combinations whose distinct count equals the row count.

    Deterministic and biased the way SAP tables are laid out: only the first
    MAX_KEY_COLUMNS non-constant, fully populated columns (excluding MANDT) are
    considered. All arities up to MAX_KEY_ARITY are searched; supersets of a
    smaller unique combination are pruned; results are ranked by the LEFT-MOST
    columns first (SAP key fields lead the table), then by fewer columns.

    Why not stop at the first unique single column: a description or timestamp
    column can be COINCIDENTALLY unique in a sample and would then mask the
    true composite key sitting to its left (found by this module's own test on
    a synthetic EQKT: EQKTX unique by luck, EQUNR+SPRAS the real key). Ranking
    by position resolves that deterministically; the residue is what the v2
    judgement layer will sanity-check."""
    row_count: int = int(frame.shape[0])
    columns: list[str] = []
    position_of: dict[str, int] = {}                                     # v1.1
    name: str = ""
    series: pd.Series = None
    unique_combos: list[list[str]] = []                                  # v1.1
    arity: int = 0
    combo: tuple = ()
    existing: list[str] = []

    if row_count == 0:
        return []

    for name in frame.columns:
        if name == "MANDT":
            continue
        series = frame[name]
        if series.isna().any():
            continue          # a key column must be fully populated
        if series.nunique() <= 1:
            continue          # constants cannot contribute to a key
        position_of[name] = len(columns)                                 # v1.1
        columns.append(name)
        if len(columns) >= MAX_KEY_COLUMNS:
            break

    for arity in range(1, MAX_KEY_ARITY + 1):                            # v1.1
        for combo in itertools.combinations(columns, arity):
            # Prune supersets of an already-found unique combination.
            if any(set(existing).issubset(set(combo)) for existing in unique_combos):
                continue                                                 # v1.1
            if int(frame[list(combo)].drop_duplicates().shape[0]) == row_count:
                unique_combos.append(list(combo))                        # v1.1

    unique_combos.sort(key=_key_rank_builder(position_of))               # v1.1
    return unique_combos


def _key_rank_builder(position_of: dict[str, int]):                      # v1.1
    """Rank a candidate by the right-most column it uses (left-most combos
    win), then by arity. Named function rather than a lambda, per house style."""
    def rank(combo: list[str]) -> tuple[int, int]:
        rightmost: int = max(position_of[name] for name in combo)
        return (rightmost, len(combo))
    return rank


# ---------------------------------------------------------------------------
# Role mapping: lookup against the controlled vocabulary
# ---------------------------------------------------------------------------

def map_roles(
    header_fields: list[str],
    roles: dict[str, dict[str, Any]],
) -> tuple[dict[str, str], list[str]]:
    """Map each field to a role via the vocabulary's sap_examples. Returns
    (mapped: field -> role_id, unmapped fields). Unmapped is not failure - the
    interpreter will characterise those fields and inference handles the rest."""
    example_index: dict[str, str] = {}
    role_id: str = ""
    entry: dict[str, Any] = {}
    example: str = ""
    mapped: dict[str, str] = {}
    unmapped: list[str] = []
    field_name: str = ""

    for role_id, entry in roles.items():
        for example in entry.get("sap_examples", []) or []:
            example_index[example] = role_id

    for field_name in header_fields:
        if field_name == "MANDT":
            continue
        if field_name in example_index:
            mapped[field_name] = example_index[field_name]
        else:
            unmapped.append(field_name)
    return mapped, unmapped


def reference_readiness(
    mapped: dict[str, str],
    roles: dict[str, dict[str, Any]],
    manifest_path: str | Path,
) -> list[dict[str, Any]]:
    """For each mapped role that validates against a reference table, report
    whether that table is loaded in the manifest."""
    raw: Any = None
    status_by_table: dict[str, str] = {}
    report: list[dict[str, Any]] = []
    field_name: str = ""
    role_id: str = ""
    reference_table: Optional[str] = None

    manifest_file = Path(manifest_path)
    if manifest_file.exists():
        raw = yaml.safe_load(manifest_file.read_text(encoding="utf-8"))
        for entry in (raw.get("tables", []) if isinstance(raw, dict) else []):
            status_by_table[str(entry.get("name"))] = str(entry.get("status", "pending_extract"))

    for field_name, role_id in sorted(mapped.items()):
        reference_table = (roles.get(role_id) or {}).get("reference_table")
        if reference_table is None:
            continue
        report.append({
            "field": field_name,
            "role": role_id,
            "reference_table": reference_table,
            "status": status_by_table.get(reference_table, "not_in_manifest"),
        })
    return report


# ---------------------------------------------------------------------------
# Scaffold
# ---------------------------------------------------------------------------

def _field_stub(field_profile: Any) -> dict[str, Any]:  # v2.0
    """A skeleton FieldSpec block seeded from the profile. Type and domain are
    PROPOSALS a scan can support; mandatory-ness and business meaning are
    judgement the steward supplies (or, later, the v2 layer proposes)."""
    type_hint: str = getattr(field_profile, "type_hint", "free_text")
    inferred_domain: Any = getattr(field_profile, "inferred_domain", None)
    populated_pct: float = float(getattr(field_profile, "populated_pct", 0.0))
    stub: dict[str, Any] = {}

    stub["description"] = "TODO"
    stub["role"] = "key" if type_hint == "key" else "attribute"
    stub["type"] = {"key": "key", "categorical": "code", "constant": "code",
                    "numeric_text": "text"}.get(type_hint, "text")
    stub["mandatory"] = populated_pct >= 99.99
    stub["length"] = {
        "min": int(getattr(field_profile, "min_length", 0)),
        "max": int(getattr(field_profile, "max_length", 0)),
    }
    stub["observed_population_pct"] = populated_pct
    if isinstance(inferred_domain, list) and inferred_domain:
        stub["domain"] = list(inferred_domain)
    return stub


def _uniqueness_stub() -> dict:  # v2.1
    """A DRAFT uniqueness block in the schema v0.4 shape.

    Every value is a starting point the steward must confirm. Blocking keys
    must be fields two records have to agree on EXACTLY before they are ever
    compared; compare fields are the text that is then scored for similarity.
    The bands and weights are the stated defaults and are not calibrated.
    """
    return {
        "scope": None,
        "blocking_keys": ["TODO"],
        "compare_fields": [{"field": "TODO", "weight": 1.0}],
        "methods": {
            "fuzzy": {"metric": "jaro_winkler", "weight": 0.5},
            "semantic": {"model": "all-MiniLM-L6-v2", "weight": 0.5},
        },
        "bands": {"duplicate": 0.92, "review_low": 0.8},
    }


def scaffold(
    xlsx_path: str | Path,
    table: str,
    roles_path: str | Path = "config/rule_bank/field_roles.yaml",
    manifest_path: str | Path = "config/reference/manifest.yaml",
    sheet: str = "Data",
) -> dict[str, Any]:
    """Run the full deterministic analysis and return a draft SCHEMA STUB plus
    the readiness report. Pure function of its inputs; the CLI does the
    printing."""
    header_row: int = 0
    header_fields: list[str] = []
    anchor: str = ""
    frame: pd.DataFrame = None
    key_candidates: list[list[str]] = []
    roles: dict[str, dict[str, Any]] = {}
    mapped: dict[str, str] = {}
    unmapped: list[str] = []
    readiness: list[dict[str, Any]] = []
    draft_schema: dict[str, Any] = {}
    profile: Any = None
    field_blocks: dict[str, Any] = {}
    name: str = ""

    header_row, header_fields = detect_header(xlsx_path, sheet=sheet)
    anchor = next((f for f in header_fields if f != "MANDT"), header_fields[0])
    frame = load_sap_table(str(xlsx_path), header_anchor=anchor, sheet=sheet)
    key_candidates = detect_key_candidates(frame)
    roles = load_field_roles(roles_path)
    mapped, unmapped = map_roles(list(frame.columns), roles)
    readiness = reference_readiness(mapped, roles, manifest_path)

    # Seed the field blocks from a real profile of the extract.
    profile = profile_table(frame, table)
    for name in frame.columns:
        field_blocks[name] = _field_stub(profile.fields[name])

    draft_schema = {
        "table": table,
        "description": f"TODO: describe the {table} object",
        "source_system": "TODO",
        "primary_key": key_candidates[0] if key_candidates else ["TODO"],
        "header_anchor": anchor,
        "file_pattern": Path(xlsx_path).name.replace(table, "{table}"),
        "uniqueness": _uniqueness_stub(),  # v2.1
        "fields": field_blocks,
    }
    return {
        "draft_schema": draft_schema,
        "header_row": header_row,
        "row_count": int(frame.shape[0]),
        "column_count": int(frame.shape[1]),
        "key_candidates": key_candidates,
        "role_mapped": mapped,
        "role_unmapped": unmapped,
        "reference_readiness": readiness,
    }


def main() -> None:
    parser: argparse.ArgumentParser = argparse.ArgumentParser(
        description="Scaffold a draft SCHEMA from a raw SE16N/SE12 extract (no LLM)."
    )
    parser.add_argument("--file", required=True, help="path to the xlsx extract")
    parser.add_argument("--table", required=True, help="SAP table name, e.g. EQUI")
    parser.add_argument("--roles", default="config/rule_bank/field_roles.yaml")
    parser.add_argument("--manifest", default="config/reference/manifest.yaml")
    parser.add_argument("--out", default=None, help="write the draft schema YAML here, e.g. config/schema/equi.yaml")
    args: argparse.Namespace = parser.parse_args()

    result: dict[str, Any] = scaffold(args.file, args.table, args.roles, args.manifest)
    draft: dict[str, Any] = result["draft_schema"]  # v2.0
    entry: dict[str, Any] = {}

    print(f"=== {args.table}: {result['row_count']} rows, {result['column_count']} columns "
          f"(header row {result['header_row']}) ===")
    print(f"key candidates : {result['key_candidates'] or 'NONE FOUND - set primary_key manually'}")
    print(f"roles mapped   : {len(result['role_mapped'])} fields")
    print(f"roles unmapped : {len(result['role_unmapped'])} fields "
          f"(these route to inference; consider new roles for the important ones)")
    for entry in result["reference_readiness"]:
        print(f"  reference {entry['reference_table']:<8} ({entry['field']} -> {entry['role']}): "
              f"{entry['status']}")

    if args.out:
        out_path = Path(args.out)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with out_path.open("w", encoding="utf-8") as handle:
            handle.write("# DRAFT schema - generated by tools/onboard_object.py\n")
            handle.write("# Review every TODO (descriptions, mandatory flags, domains),\n")
            handle.write("# then confirm by removing this banner.\n")
            yaml.safe_dump(draft, handle, sort_keys=False, allow_unicode=True)
        print(f"\ndraft schema written: {out_path} (review the TODOs before use)")
    else:
        print("\n--- draft schema (pass --out to write it) ---")
        print(yaml.safe_dump(draft, sort_keys=False, allow_unicode=True))


if __name__ == "__main__":
    main()
